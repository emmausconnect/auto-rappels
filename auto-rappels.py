import datetime
import openpyxl
import re
import sys
import time

from ipaddress import IPv4Address
from functools import partial
from pathlib import Path
from pyairmore.request import AirmoreSession
from pyairmore.services.messaging import MessagingService
from PyQt5.QtCore import QThread, pyqtSignal, Qt, QDir
from PyQt5.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QFileDialog,
    QMessageBox,
    QProgressDialog,
    QPushButton,
    QApplication,
    QLineEdit,
    QStyledItemDelegate,
    QTableWidget,
    QTableWidgetItem,
    QCheckBox,
    QTextEdit,
)
from PyQt5.QtGui import (
    QPixmap,
    QFont,
    QFontDatabase,
    QTextCursor,
    QColor,
    QTextCharFormat,
)

db_path = Path("./bin/database/db.xlsx")
bg_img = "./bin/img/BG.jpg"
font_path = "Fira Sans/FiraSans-Regular.ttf"
font = "Fira Sans"
stylesheet = Path("./style.qss")


def is_valid_hour(hour):
    """Vérifie si l'heure est valide"""

    regex = r"^([01][0-9]|2[0-3]):[0-5][0-9]$"
    return bool(re.match(regex, hour))


def is_valid_phone_number(phone_number):
    """Vérifie si le numéro de téléphone est valide"""

    regex = r"^(?:\D*\d){0}\D*0(?:\D*\d){9}\D*$"
    return bool(re.match(regex, phone_number))


def is_valid_date(date):
    """Vérifier si la date est valide"""

    regex = r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$"
    return bool(re.match(regex, date))


class AirmoreConnectionThread(QThread):
    """
    Connexion à Airmore et d'envoi du SMS
    La connexion s'effectue dans un thread séparé afin de ne pas bloquer l'application pendant l'attente de connexion

    Attributs
    ----------

    Méthodes
    -------
    """

    connection_success = pyqtSignal()
    connection_error = pyqtSignal(str)

    def __init__(self, ip_address, chemin):
        """
        Construit l'objet

        Paramètres
        ----------
            ip_address: str
                adresse IP du smartphone
            chemin: str
                chemin vers le fichier .xlsx
            session: AirmoreSession

        """
        super().__init__()
        self.chemin = chemin
        self.db = openpyxl.load_workbook(db_path)
        self.file_data = openpyxl.load_workbook(chemin)
        self.session = AirmoreSession(IPv4Address(ip_address))

    def run(self):
        """Connexion au téléphone via ip et pyAiremore"""
        try:
            service = MessagingService(self.session)

            while not self.session.is_server_running:
                time.sleep(1)
            if self.session.request_authorization() == True:
                self.connection_success.emit()
                self.sheet = self.file_data.active
                for row_index in range(1, self.sheet.max_row + 1):

                    if (
                        self.sheet.cell(row=row_index, column=26).value == 1  # permco
                        or self.sheet.cell(row=row_index, column=26).value
                        == 2  # parcours
                        or self.sheet.cell(row=row_index, column=26).value == 3  # RDV
                    ):
                        service.send_message(
                            str(
                                self.sheet.cell(row=row_index, column=5).value
                            ),  # envoyer le message service.send_message(numéro inscrit colonne 5 du fichier xlxs,
                            self.replace_message(
                                self.db.active.cell(
                                    int(
                                        self.sheet.cell(row=row_index, column=26).value
                                    ),
                                    1,
                                ).value,
                                row_index,
                            ),
                        )  #                                                 """"
                        self.sheet.cell(row=row_index, column=26).value = (
                            0  # remise à zero de la cellule z
                        )
                self.file_data.save(self.chemin)
            else:
                QMessageBox.critical(self, "Erreur", "Connexion refusé")
        except Exception as e:
            self.connection_error.emit(str(e))

    def replace_message(self, message, row_index):
        """Remplace les keywords dans le message"""
        days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        months = [
            "Janvier",
            "Février",
            "Mars",
            "Avril",
            "Mai",
            "Juin",
            "Juillet",
            "Août",
            "Septembre",
            "Octobre",
            "Novembre",
            "Décembre",
        ]
        keywords = {
            "[heure]": str(self.sheet.cell(row=row_index, column=2).value)[11:],
            "[date]": str(self.sheet.cell(row=row_index, column=2).value)[:10],
            "[nom]": str(self.sheet.cell(row=row_index, column=4).value).split(" ")[0],
            "[prenom]": str(self.sheet.cell(row=row_index, column=4).value).split(" ")[
                -1
            ],
            "[jour_semaine]": days[
                datetime.datetime.strptime(
                    str(self.sheet.cell(row=row_index, column=2).value)[:10], "%d/%m/%Y"
                ).weekday()
            ],
            "[mois]": months[
                datetime.datetime.strptime(
                    str(self.sheet.cell(row=row_index, column=2).value)[:10], "%d/%m/%Y"
                ).month
                - 1
            ],
            "[annee]": str(self.sheet.cell(row=row_index, column=2).value)[6:10],
            "[tel]": str(self.sheet.cell(row=row_index, column=5).value),
            "[jour_num]": str(self.sheet.cell(row=row_index, column=2).value)[:2],
        }
        for key, value in keywords.items():
            message = message.replace(key, value)
        return message


class FontDelegate(QStyledItemDelegate):
    """
    Ajoute une police lors de la modification dans les tableaux
    """

    def createEditor(self, parent, option, index):
        editor = QLineEdit(parent)
        editor.setFont(QFont(font, 12))
        return editor


class Button(QPushButton):
    def __init__(self, title, container, slot, x, y, width=200, height=50):
        """
        Paramètres
        ----------
            title: str
                titre du bouton
            container: QWidget
                container
            x: int
                abcisse du coin supérieur gauche
            y: int
                ordonnée du coin supérieur gauche
            width:
                largeur
            height:
                hauteur
            slot:
                fonction appelée lorsqu'on clique sur le bouton
        """
        super().__init__(title, container)
        self.setGeometry(x, y, width, height)
        self.clicked.connect(slot)


class MainWindow(QWidget):
    """
    Gestion de tous les éléments de logique de la fenêtre principale

    Attributs
    ---------

    Méthodes
    --------
    """

    def __init__(self, width=1200, height=850):
        """
        Construit l'objet fenêtre principale. Pour s'assurer que les éléments restent au centre de la fenêtre, il y a 2 grilles de placement : une verticale (layout), et une horizontale (layout_h)
        """

        super().__init__()
        self.db = openpyxl.load_workbook(db_path)
        self.selected_file = None
        self.highlighting = False
        self.setWindowTitle("AutoRappel V2")
        self.setMinimumSize(width, height)
        self.container = QWidget(self)
        self.container.setGeometry(0, 0, width, height)
        self.container.setMinimumSize(width, height)

        layout = QVBoxLayout(self)
        layout.setAlignment(Qt.AlignCenter)

        layout_h = QHBoxLayout()
        layout_h.setAlignment(Qt.AlignCenter)

        layout.addLayout(layout_h)
        layout_h.addWidget(self.container)

        self.set_bg()

        # self.setStyleSheet(stylesheet.read_text())
        self.show_buttons()

    def set_bg(self):
        """
        Définition de l'image EC en haut à gauche de la fenêtre
        Appelée après chaque clean des widgets
        """

        background_label = QLabel(self.container)
        pixmap = QPixmap(bg_img)
        background_label.setPixmap(pixmap)
        background_label.setGeometry(0, 0, self.width(), 100)
        background_label.show()

    def connexion(self, type_num):
        """
        Paramètres
        ----------
            type_num: str
                permco | rdv | parcours

        """

        self.clear_widgets()
        self.set_bg()
        self.data = openpyxl.load_workbook(db_path)

        self.ip_input = QLineEdit(self.container)
        self.ip_input.setPlaceholderText("Entrez l'adresse IP du téléphone")
        self.ip_input.setStyleSheet("background-color: white;")
        self.ip_input.setGeometry(350, 250, 500, 50)
        self.ip_input.setFont(QFont(font, 14))
        last_ip = self.data.active.cell(row=4, column=1).value
        self.ip_input.setText(last_ip)
        self.ip_input.setAlignment(Qt.AlignCenter)
        self.ip_input.show()

        send_button = Button(
            "Envoyer SMS", self.container, self.connect_and_send_sms, 500, 350
        )
        send_button.show()

        quit_button = Button("Quitter", self.container, self.quit, 950, 25)
        quit_button.show()

        back_button = Button(
            "Retour", self.container, 725, 25, partial(self.mess, type_num)
        )
        back_button.show()

    def quit(self):
        QApplication.instance().quit()

    def connect_and_send_sms(self):
        """Lance le thread airmore en lui passant les arguments ip et type_num"""
        ip_address = self.ip_input.text()
        sheet = self.selected_file
        self.data.active.cell(row=4, column=1).value = ip_address
        self.data.save(db_path)

        if not ip_address:
            QMessageBox.critical(self, "Erreur", "Veuillez entrer une adresse IP.")
            return

        self.connection_thread = AirmoreConnectionThread(ip_address, sheet)
        self.connection_thread.connection_success.connect(
            self.show_connection_success_message
        )
        self.connection_thread.connection_error.connect(
            self.show_connection_error_message
        )
        self.connection_thread.start()

        self.progress_dialog = QProgressDialog(
            "Accepter la connexion sur le téléphone...", "Annuler", 0, 0
        )
        self.progress_dialog.setWindowModality(Qt.WindowModal)
        self.progress_dialog.setWindowTitle("Connexion en cours")
        # self.progress_dialog.setCancelButton(None)
        self.progress_dialog.show()

    def show_connection_success_message(self):
        """
        Message affiché en cas de succès de la connexion
        """
        self.progress_dialog.close()
        self.clear_widgets()

        msg = QMessageBox()
        msg.setText(
            "La connexion avec le téléphone a été établie avec succès. Tous les messages seront envoyés"
        )
        msg.setWindowTitle("Connexion réussie")
        msg.setIcon(QMessageBox.Information)
        msg.setStyleSheet("background-color: rgb(255, 255, 255)")
        msg.exec_()

        self.show_buttons()

    def show_connection_error_message(self, error_message):
        self.progress_dialog.close()

        msg = QMessageBox()
        msg.critical(
            self.container,
            "Erreur",
            f"Impossible de se connecter avec l'adresse IP {self.ip_input.text()}: {error_message}",
        )
        msg.setStyleSheet("QWidget { background-color: white; }")

    def clear_widgets(self):
        """
        Supprime tous les widgets présents dans le container pour permettre l'affichage d'autres widgets
        """
        for sub_widget in self.container.findChildren(QWidget):
            sub_widget.deleteLater()

    def browse_file(self):
        """
        Parcours un fichier .xslx
        """

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Sélectionner un fichier", "", "Excel (*.xlsx)"
        )

        if file_path:
            self.selected_file = file_path

            msg = QMessageBox()
            msg.setText(f"Le fichier sélectionné est : {file_path}")
            msg.setWindowTitle("Fichier sélectionné")
            msg.setIcon(QMessageBox.Information)
            msg.setStyleSheet("background-color: rgb(255, 255, 255)")
            msg.exec_()

            self.file_input.setText(file_path)
            self.file_input.setReadOnly(False)

    def show_buttons(self):
        """
        Fonction principale de sélection de fichier et choix du type de rappel
        """
        self.clear_widgets()
        self.set_bg()

        permco_button = Button(
            "Permanences connectées", self.container, partial(self.handle, 1), 500, 350
        )
        permco_button.show()

        self.file_input = QLineEdit(self.container)
        self.file_input.setPlaceholderText("Chemin du rapport excel")
        self.file_input.setStyleSheet("background-color: white;")
        self.file_input.setGeometry(350, 250, 500, 50)
        self.file_input.setFont(QFont(font, 14))
        self.file_input.setReadOnly(True)
        self.file_input.show()

        try:
            self.file_input.setText(self.selected_file)
            self.file_input.setReadOnly(False)
            for row_index in range(1, self.file_data.active.max_row + 1):
                self.file_data.active.cell(row=row_index, column=26).value = (
                    0  # remise a zero de la cellule z
                )
            self.file_data.save(self.selected_file)
        except Exception as e:
            print("Aucun fichier chargé")
            print(e)

        parcours_button = Button(
            "Parcours", self.container, partial(self.handle, 2), 500, 425
        )
        parcours_button.show()

        rendezvous_button = Button(
            "Rendez-vous", self.container, partial(self.handle, 3), 500, 500
        )
        rendezvous_button.show()

        browse_button = Button(
            "Choisir un fichier", self.container, self.browse_file, 350, 150, 500, 50
        )
        browse_button.show()

        quit_button = Button("Rendez-vous", self.container, self.quit, 950, 25)
        quit_button.show()

    def is_file_valid(self, file_path):
        """
        Ouvre le fichier et gère les erreurs
        """

        try:
            wb = openpyxl.load_workbook(file_path)
            self.file_data = wb
            return True

        except Exception as e:
            self.file_data = None
            return e

    def handle(self, type_num):
        """
        Affiche les données du fichier xlsx sous forme de tableau. Permet de sélectionner les contacts pour l'envoi du sms.
        """

        if not self.selected_file:
            QMessageBox.warning(
                self, "Avertissement", "Veuillez d'abord sélectionner un fichier."
            )
            return

        if self.is_file_valid(self.selected_file) != True:
            QMessageBox.warning(
                self,
                "Avertissement",
                f"Impossible de charger le fichier : {self.is_file_valid(self.selected_file)}",
            )
            return

        sheet = self.file_data.active
        num_rows = sheet.max_row
        displayed_row_index = 0
        num_displayed_rows = 0

        type_text = None
        if type_num == 1:
            type_text = "Permanence Connectée"
        elif type_num == 2:
            type_text = "Parcours d'initiation"
        elif type_num == 3:
            type_text = "RDV bénéficiaire"

        try:
            for row_index in range(1, num_rows + 1):
                type_rdv = sheet.cell(row=row_index, column=10).value
                if type_text in type_rdv:
                    num_displayed_rows += 1
        except:
            QMessageBox.warning(
                self,
                "Avertissement",
                f"Impossible de charger le fichier : {self.selected_file}",
            )
            return

        self.clear_widgets()
        self.set_bg()

        self.table_widget = QTableWidget(self.container)
        self.table_widget.setGeometry(40, 100, 1120, 650)

        headers = [
            "Nom complet",
            "Date",
            "Heure",
            "Numéro Tel",
            "Observations",
            "Statut",
            "Rappel",
        ]

        self.table_widget.setColumnCount(len(headers))
        self.table_widget.setHorizontalHeaderLabels(headers)

        self.table_widget.setColumnWidth(0, 200)
        self.table_widget.setColumnWidth(1, 150)
        self.table_widget.setColumnWidth(2, 100)
        self.table_widget.setColumnWidth(3, 200)
        self.table_widget.setColumnWidth(4, 200)
        self.table_widget.setColumnWidth(5, 120)
        self.table_widget.setColumnWidth(6, 90)

        self.table_widget.setRowCount(num_displayed_rows)

        for row_index in range(1, num_rows + 1):
            type_rdv = sheet.cell(row=row_index, column=10).value
            if type_text in type_rdv:

                nom_complet = QTableWidgetItem(
                    str(sheet.cell(row=row_index, column=4).value)
                )
                date = str(sheet.cell(row=row_index, column=2).value)[:10]
                heure = str(sheet.cell(row=row_index, column=2).value)[11:]
                numero_tel = str(sheet.cell(row=row_index, column=5).value)
                statut_z = sheet.cell(row=row_index, column=26).value
                statut_participant = QTableWidgetItem(
                    str(sheet.cell(row=row_index, column=7).value)
                )
                if type_num == 3:  # RDV
                    if str(sheet.cell(row=row_index, column=8).value) == "None":
                        observation = QTableWidgetItem(
                            str(sheet.cell(row=row_index, column=1).value)
                        )
                    else:
                        observation = QTableWidgetItem(
                            str(sheet.cell(row=row_index, column=1).value)
                            + ": "
                            + str(sheet.cell(row=row_index, column=8).value)
                        )
                else:
                    if str(sheet.cell(row=row_index, column=8).value) == "None":
                        observation = QTableWidgetItem("")
                    else:
                        observation = QTableWidgetItem(
                            str(sheet.cell(row=row_index, column=8).value)
                        )

                self.table_widget.setItem(displayed_row_index, 0, nom_complet)
                nom_complet.setTextAlignment(Qt.AlignCenter)
                nom_complet.setFont(QFont(font, 12))

                self.table_widget.setItem(displayed_row_index, 4, observation)
                observation.setFont(QFont(font, 12))

                self.table_widget.setItem(displayed_row_index, 5, statut_participant)
                statut_participant.setTextAlignment(Qt.AlignCenter)
                statut_participant.setFont(QFont(font, 12))

                checkbox_widget = QWidget()
                checkbox_layout = QHBoxLayout()
                checkbox = QCheckBox()
                checkbox.setChecked(bool(statut_z))
                checkbox.setProperty("id", str(row_index))
                checkbox_layout.addWidget(checkbox)
                checkbox_layout.setAlignment(Qt.AlignCenter)
                checkbox_widget.setLayout(checkbox_layout)
                self.table_widget.setCellWidget(displayed_row_index, 6, checkbox_widget)

                date_widget = QTableWidgetItem(str(date))
                if not is_valid_date(date):
                    date_widget.setForeground(QColor(Qt.red))
                    checkbox_widget.setEnabled(False)
                date_widget.setTextAlignment(Qt.AlignCenter)
                date_widget.setFont(QFont(font, 12))
                self.table_widget.setItem(displayed_row_index, 1, date_widget)

                hour_widget = QTableWidgetItem(str(heure))
                if not is_valid_hour(heure):
                    hour_widget.setForeground(QColor(Qt.red))
                    checkbox_widget.setEnabled(False)
                hour_widget.setTextAlignment(Qt.AlignCenter)
                hour_widget.setFont(QFont(font, 12))
                self.table_widget.setItem(displayed_row_index, 2, hour_widget)

                phone_widget = QTableWidgetItem(numero_tel)
                if not is_valid_phone_number(numero_tel):
                    phone_widget.setForeground(QColor(Qt.red))
                    checkbox_widget.setEnabled(False)
                phone_widget.setTextAlignment(Qt.AlignCenter)
                phone_widget.setFont(QFont(font, 12))
                self.table_widget.setItem(displayed_row_index, 3, phone_widget)

                self.table_widget.setItemDelegate(FontDelegate())
                displayed_row_index += 1

        self.table_widget.cellChanged.connect(self.set_enabled)

        global_checkbox = QCheckBox("Sélectionner/Désélectionner Tout", self.container)
        global_checkbox.setGeometry(50, 750, 280, 50)
        global_checkbox.setStyleSheet("color: white;")
        global_checkbox.setChecked(False)
        global_checkbox.stateChanged.connect(self.check_uncheck_all)
        global_checkbox.show()

        self.table_widget.resizeRowsToContents()

        validate_button = Button(
            "Valider",
            self.container,
            partial(self.save_checked_values, type_num),
            500,
            760,
        )
        validate_button.show()

        quit_button = Button("Quitter", self.container, self.quit, 950, 25)
        quit_button.show()

        back_button = Button("Retour", self.container, self.show_buttons, 725, 25)
        back_button.show()

        self.table_widget.show()

    def check_uncheck_all(self, state):
        """
        Cocher ou décocher toutes les cases à cocher de la colonne en fonction de l'état de la case à cocher de l'en-tête
        """

        for row_index in range(self.table_widget.rowCount()):
            checkbox_widget = self.table_widget.cellWidget(row_index, 6)
            if (
                self.check_phone_number(row_index, 3)
                and self.check_date(row_index, 1)
                and self.check_hour(row_index, 2)
            ):
                if checkbox_widget is not None:
                    checkbox = checkbox_widget.findChild(QCheckBox)
                    if checkbox is not None:
                        checkbox.setChecked(state == Qt.Checked)

    def update_status_z(self, state, row, row_xls, type_num):
        """
        Met à jour l'état de la colonne Z en fonction de l'état de la case à cocher

        Paramètres
        ----------
            state:
            row: int
                index de la ligne du tableau
            row_xls: int
                index de la ligne fichier
            type_num:
                type de RDV
        """

        sheet = self.file_data.active

        if int(state) == 1:  # si case cocher
            sheet.cell(row=row_xls, column=26).value = int(
                type_num
            )  # remplit la colonne z par le type de rdv (format num)
        else:
            sheet.cell(row=row_xls, column=26).value = int(
                state
            )  # sinon met la colonne z à 0

            # remplit le fichier par les valeurs modifiées du tableau
        sheet.cell(row=row_xls, column=4).value = self.table_widget.item(
            row, 0
        ).text()  #                       ""
        sheet.cell(row=row_xls, column=2).value = (
            self.table_widget.item(row, 1).text()
            + " "
            + self.table_widget.item(row, 2).text()
        )
        sheet.cell(row=row_xls, column=5).value = self.table_widget.item(row, 3).text()
        sheet.cell(row=row_xls, column=8).value = self.table_widget.item(row, 4).text()
        sheet.cell(row=row_xls, column=7).value = self.table_widget.item(row, 5).text()

        self.file_data.save(self.selected_file)

    def save_checked_values(self, type_num):
        """
        Enregistre les modifications et le compte du nombre de cases cochées
        """

        if self.file_data is None:
            QMessageBox.warning(
                self,
                "Avertissement",
                "Veuillez d'abord sélectionner un fichier.",
            )
            return

        messages_to_send = 0

        for row_index in range(self.table_widget.rowCount()):
            checkbox_widget = self.table_widget.cellWidget(row_index, 6)
            if checkbox_widget is not None:
                checkbox = checkbox_widget.findChild(QCheckBox)
                if checkbox is not None:
                    checked = checkbox.isChecked()
                    if checked:
                        messages_to_send += 1

                    self.update_status_z(
                        checked,
                        row_index,
                        int(checkbox.property("id")),
                        type_num,
                    )

        if messages_to_send > 0:
            end_message = f"{messages_to_send} messages seront envoyés"
            if messages_to_send == 1:
                end_message = "1 message sera envoyé"

            QMessageBox.information(self, "Succès", end_message)
            self.mess(type_num)
        else:
            QMessageBox.warning(
                self,
                "Avertissement",
                "Veuillez sélectionner des contacts",  # sinon afficher message d'erreur pour demander de sélectionné des contacts
            )

    def check_phone_number(self, row, column):

        phone_widget = self.table_widget.item(row, 3)

        checkbox_widget = self.table_widget.cellWidget(phone_widget.row(), 6)
        checkbox = checkbox_widget.findChild(QCheckBox)

        phone_number = phone_widget.text()

        if not is_valid_phone_number(phone_number):
            phone_widget.setForeground(QColor(Qt.red))
            checkbox.setChecked(False)
            checkbox_widget.setEnabled(False)
            return False
        else:
            phone_widget.setForeground(QColor(Qt.black))
            return True

    def check_date(self, row, column):

        date_widget = self.table_widget.item(row, 1)

        checkbox_widget = self.table_widget.cellWidget(date_widget.row(), 6)
        checkbox = checkbox_widget.findChild(QCheckBox)

        date = date_widget.text()

        if not self.is_valid_date(date):
            date_widget.setForeground(QColor(Qt.red))
            checkbox.setChecked(False)
            checkbox_widget.setEnabled(False)
            return False
        else:
            date_widget.setForeground(QColor(Qt.black))
            return True

    def check_hour(self, row, column):

        hour_widget = self.table_widget.item(row, 2)

        checkbox_widget = self.table_widget.cellWidget(hour_widget.row(), 6)
        checkbox = checkbox_widget.findChild(QCheckBox)

        hour = hour_widget.text()

        if not self.is_valid_hour(hour):
            hour_widget.setForeground(QColor(Qt.red))
            checkbox.setChecked(False)
            checkbox_widget.setEnabled(False)
            return False
        else:
            hour_widget.setForeground(QColor(Qt.black))
            return True

    def set_enabled(self, row, column):
        """
        Teste si les formats sont correctes avant d'activer la case à cocher par défaut
        """

        checkbox_widget = self.table_widget.cellWidget(row, 6)

        if (
            self.check_phone_number(row, column)
            and self.check_date(row, column)
            and self.check_hour(row, column)
        ):
            checkbox_widget.setEnabled(True)
        else:
            self.check_date(row, column)
            self.check_hour(row, column)

    def mess(self, type_num):
        """
        Permet d'éditer le message avant l'envoi
        """

        self.clear_widgets()
        self.set_bg()

        quit_button = Button("Quitter", self.container, self.quit, 950, 25)
        quit_button.show()

        back_button = Button(
            "Retour", self.container, partial(self.handle, type_num), 725, 25
        )
        back_button.show()

        send_button = Button(
            "Envoyer", self.container, partial(self.save_db_changed, type_num), 500, 700
        )
        send_button.show()

        self.text_edit = QTextEdit(self.container)
        self.text_edit.setPlaceholderText("Entrez votre message ici...")
        self.text_edit.setText(str(self.db.active.cell(type_num, 1).value))
        self.text_edit.setGeometry(300, 150, 600, 450)
        self.text_edit.setFont(QFont(font, 20))
        self.text_edit.textChanged.connect(self.onTextChanged)
        self.highlight_keywords()
        self.text_edit.show()

    def onTextChanged(self):
        """
        Temporise pour ne pas surcharger d'exécution la fonction de surbrillance
        """
        if not self.highlighting:
            self.highlighting = True
            self.highlight_keywords()
            self.highlighting = False

    def save_db_changed(self, type_num):
        self.db.active.cell(1, 1).value = self.text_edit.toPlainText()
        self.db.save(db_path)
        self.connexion(type_num)

    def highlight_keywords(self, color="#00acb0"):
        """
        Met en surbrillance les mots-clefs
        """

        text = self.text_edit.toPlainText()
        keywords = [
            "[heure]",
            "[date]",
            "[nom]",
            "[prenom]",
            "[tel]",
            "[jour_semaine]",
            "[mois]",
            "[annee]",
            "[jour_num]",
        ]

        red_format = QTextCharFormat()
        red_format.setForeground(QColor(color))

        cursor = self.text_edit.textCursor()
        cursor.select(QTextCursor.Document)
        cursor.setCharFormat(QTextCharFormat())

        for keyword in keywords:
            index = text.find(keyword)
            while index != -1:
                cursor.setPosition(index)
                cursor.movePosition(
                    QTextCursor.Right, QTextCursor.KeepAnchor, len(keyword)
                )
                cursor.setCharFormat(red_format)
                index = text.find(keyword, index + len(keyword))


if __name__ == "__main__":
    app = QApplication(sys.argv)
    dir_ = QDir(font)
    QFontDatabase.addApplicationFont(font_path)
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())
