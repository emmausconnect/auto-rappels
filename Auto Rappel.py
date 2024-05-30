import sys
from PyQt5.QtWidgets import *
from PyQt5.QtGui import *
from PyQt5.QtCore import *
import time
import openpyxl
import re
import datetime
from pyairmore.request import AirmoreSession
from pyairmore.services.messaging import MessagingService
from ipaddress import IPv4Address
from functools import partial


# class de connexion à airemore et d'envoi du SMS
# La connexion s'effectue dans un thread séparé afin de ne pas bloquer l'application pendant l'attente de connexion
class AirmoreConnectionThread(QThread):
    connection_success = pyqtSignal()  # signal de succès de connexion
    connection_error = pyqtSignal(str)  # signal d'echec de connexion signal(erreur)
    session = None  # définition de la variable de stockage de la session Airmore

    def __init__(
        self, ip_address, chemin, parent=None
    ):  # init(self, adresse_ip, chemin du fichier .xlsx, None)
        super().__init__(parent)
        self.ip_address = IPv4Address(
            ip_address
        )  # recupération de l'adresse ip et passage au format "IP"
        self.chemin = chemin  # ouverture de la variable du chemin vers le fichier .xlsx au reste de la class
        wb = openpyxl.load_workbook(
            chemin
        )  # ouverture du fichier .xlsx via openpyXl et stockage du fichier dans un variable wb
        data = openpyxl.load_workbook(
            "./database/db.xlsx"
        )  # ouverture de la base de données via openpyXl et stockage du fichier dans une variable db
        self.file_data = wb  # ouverture au reste de la class
        self.db = data  #                '''


    def run(self):  # connexion au téléphone via ip et pyAiremore run(self)
        try:
            self.session = AirmoreSession(
                self.ip_address
            )  # stockage de la session dans la variable session
            was_accepted = (
                self.session.request_authorization()
            )  # stockage de la réponse utilisateur à la connexion : Bool
            service = MessagingService(
                self.session
            )  # création de la session de messaging enfant de la session global

            # Attendre que le serveur Airmore soit en cours d'exécution sur le téléphone
            while not self.session.is_server_running:
                time.sleep(1)
            if was_accepted == True:  # Si accepter
                self.connection_success.emit()  # emettre signal de succès
                self.sheet = (
                    self.file_data.active
                )  # ouverture de la feuille active du fichier .xlsx dans la variable sheet
                for row_index in range(
                    1, self.sheet.max_row + 1
                ):  # parcourir les lignes du tableau excel

                    if (
                        self.sheet.cell(row=row_index, column=26).value
                        == 1  # Si valeur de la colonne z = 1 soit permco
                        or self.sheet.cell(row=row_index, column=26).value
                        == 2  # ou Si valeur de la colonne z = 2 soit parcours
                        or self.sheet.cell(row=row_index, column=26).value
                        == 3  # ou Si valeur de la colonne z = 3 soit rendez-vous
                    ):
                        service.send_message(  #                                             """"
                            str(self.sheet.cell(row=row_index, column=5).value), # envoyer le message service.send_message(numéro inscrit colonne 5 du fichier xlxs,
                            self.replace_mess( # retour de la fonction replace_mess(message stocké dans la base de donnés à la ligne =valeur de la colonne z de la ligne lus))
                                self.db.active.cell(  
                                    int(
                                        self.sheet.cell(row=row_index, column=26).value
                                    ),
                                    1,
                                ).value,
                                row_index,
                            ),
                        )  #                                                 """"
                        self.sheet.cell(row=row_index, column=26).value = (
                            0  # remise à zero de la cellule z
                        )
                self.file_data.save(
                    self.chemin
                )  # enregistrement des modifications dans la feuille
            else:
                QMessageBox.critical(
                    self, "Erreur", "Connexion refusé"
                )  # message en cas de refus de connexion
        except Exception as e:
            # En cas d'erreur, émettre le signal d'erreur avec le message

            self.connection_error.emit(str(e))

    def replace_mess(
        self, mess, row_index
    ):  # fonction de remplacement des keywords dans le message   replace_mess(self, message, index de la ligne xlsx)
        days = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]
        months = ["Janvier", "Février", "Mars", "Avril", "Mai", "Juin", "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"]
        keywords = {
            "[heure]": str(self.sheet.cell(row=row_index, column=2).value)[
                11:
            ],  #                                                                              """"""
            "[date]": str(self.sheet.cell(row=row_index, column=2).value)[:10],
            "[nom]": str(self.sheet.cell(row=row_index, column=4).value).split(" ")[
                0
            ],                                                                             # definion des keywords et des valeurs associées
            "[prenom]": str(self.sheet.cell(row=row_index, column=4).value).split(" ")[
                -1
            ],
            "[jour_semaine]" : days[datetime.datetime.strptime(str(self.sheet.cell(row=row_index, column=2).value)[:10], "%d/%m/%Y").weekday()
            ],
            "[mois]" : months[datetime.datetime.strptime(str(self.sheet.cell(row=row_index, column=2).value)[:10], "%d/%m/%Y").month -1
            ],
            "[annee]" : str(self.sheet.cell(row=row_index, column=2).value)[6:10],
            "[tel]" :  str(self.sheet.cell(row=row_index, column=5).value),
             "[jour_num]" : str(self.sheet.cell(row=row_index, column=2).value)[:2] #                                                                               """"""
        }
        for key, value in keywords.items():  # parcourir la liste de keywords
            mess = mess.replace(
                key, value
            )  # remplacer les occurences des keywords dans le message
        return mess  # retourner le message modifié avec les keywords remplacés

# class permettant l'ajout d'une police lors de la modification dans les tableaux
class FontDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):  #    """"
        editor = QLineEdit(parent)  # définition de la police Fira sans
        editor.setFont(QFont("Fira Sans", 12))
        return editor  #    """"


# class principal de l'app : gestion de tout les élément et logique de la fenêtre principal
class MainWindow(QWidget):

    def set_bg(self):  # definition de l'image emmaus connect en haut à gauche de la fenêtre , fonction appelé après chaque clean des widgets

        background_label = QLabel(
            self.container
        )  # création du label photo avec pour parent le container
        pixmap = QPixmap(  # définition du chemin de l'image type pixmap
            "./img/BG.jpg"  # Assurez-vous que le chemin de l'image de fond est correct
        )
        background_label.setPixmap(pixmap)  # ajout du pixmap au label photo
        background_label.setGeometry(
            0, 0, self.width(), 100
        )  # position du label dans la page ( haut à gauche du container)
        background_label.show()  # afficher l'élément

    def __init__(self):
        super().__init__()

        self.db = openpyxl.load_workbook(
            "./database/db.xlsx"
        )  # ouverture de la base de donné via openpyxl

        self.highlighting = False  # supression de la surbrillance

        self.setWindowTitle("AutoRappel V2")  # titre de la fenêtre

        self.setMinimumSize(1200, 850)  # taille minimum de la fenêtre

        self.container = QWidget(self)  # création du container
        self.container.setGeometry(
            0, 0, 1200, 850
        )  # placement dans la fenêtre principal
        self.container.setMinimumSize(1200, 850)  # taille minimum du container

        layout = QVBoxLayout(
            self
        )  # création d'une grille de placment vertical dans la fenêtre principal
        layout.setAlignment(
            Qt.AlignCenter
        )  # fait que les élément de la grille vertical reste toujours au centre de la fenêtre principal

        layout_h = QHBoxLayout()  # création d'un grille de placement horizontal
        layout_h.setAlignment(
            Qt.AlignCenter
        )  # fait que les élément de la grille vertical reste toujours au centre de la fenêtre principal

        layout.addLayout(
            layout_h
        )  # ajout de la grille horizontal dans la grille vertical
        layout_h.addWidget(
            self.container
        )  # ajout du container dans le layout horizontal

        self.set_bg()  # ajout de l'image de fond

        # feuille de style QCss (Css) pour chaque élément de la fenêtre ( police, effet, animation, couleur)
        self.setStyleSheet(
        ""    
        "   * {"
        "       background-color : #272350;"
        "        font-family: Fira Sans;"
        "    }"
        "    QMessageBox *{"
        "        background-color: rgb(255,255,255);"
        "        color : rgb(255,255,255)"
        "    }"
        "    QMessageBox QLabel, QPushButton  {"
        "        color : rgb(0,0,0)"
        "    }"
        "    QMessageBox {"
        "        background-color: rgb(255,255,255);"
        "    }"
        "    Qwidget .QPushButton:hover {"
        "    background-color: rgb(0,0,0); /* Couleur de fond au survol */"
        "    border-color: #45a049; /* Couleur de la bordure au survol */"
        "    }"
        "    QPushButton {"
        "        background-color: rgb(230,230,230);"
        "        border-style: outset;"
        "        border-width: 2px;"
        "        border-radius: 5px;"
        "        border-color: rgb(250,250,250);"
        "        font: 24px;"
        "        padding: 2px;"
        "    }"
        "    QPushButton:hover {"
        "        background-color: rgb(255, 255, 255);"
        "        border-style: inset;"
        "    }"
        "    QCheckBox {"
        "        spacing: 5px;"
        "    }"
        ""
        "    QCheckBox::indicator {"
        "        width: 20px;"
        "        height: 20px;"
        "    }"
        ""
        "    QCheckBox::indicator:unchecked {"
        "        image: url(./img/checkbox_unchecked.png);"
        "    }"
        ""
        "    QCheckBox::indicator:unchecked:hover {"
        "        image: url(./img/checkbox_hover.png);"
        "    }"
        "    QCheckBox::indicator:checked {"
        "        image: url(./img/checkbox_checked.png);"
        "    }"
        "    QCheckBox::indicator:checked:hover {"
        "        image: url(./img/checkbox_checked_hover.png);"
        "    }"
        "    QCheckBox::indicator:indeterminate:hover {"
        "        image: url(./img/checkbox_hover.png);"
        "    } "
        "    QCheckBox::indicator:disabled {"
        "        image: url(./img/checkbox_off.png); "
        "    }"
        "    "
        "   QTableWidget *{"
        "        background-color:white;"
        "        color:black;"
        "    }"
        "    QTableWidget {"
        "        background-color:white;"
        "        color:black;"
        "        text-align: center;"
        "        border:5px solid black;"
        "    }"
        "    QTableWidget::item {"
        "        text-align: center;"
        "    }"
        ""
        "    QHeaderView::section:vertical"
        "    {"
        "        background-color:white;"
        "        color: black;"
        "        border-left: 1px solid black;"
        "        border-right: 1px solid black;"
        "        border-bottom: 2px solid black;"
        "        font-family: 'Fira Sans';"
        "        "
        "        padding-left: 4px;"
        "    }"
        "    QHeaderView::section:horizontal"
        "    {"
        "        background-color:white;"
        "        color: black;"
        "        border-left: : 1px solid black;"
        "        border-right: : 1px solid black;"
        "        border-bottom : 1px solid black;"
        "        border-left: 1px dashed black;"
        "        border-right: 1px dashed black;"
        "        font-family: 'Fira Sans';"
        "        font: 24px;"
        "        padding-left: 4px;"
        "    }"
        ""
        "    QScrollBar:vertical {"
        "        border:none;"
        "        background: grey;"
        "        width: 15px;"
        "    }"
        ""
        "    QScrollBar::handle:vertical {"
        "        background: grey;"
        "        min-height: 20px;"
        "    }"
        ""
        "    QTableWidget::item {"
        "        border-left:1px dashed black;"
        "        border-bottom: 2px solid black;"
        "        font-family: 'Fira Sans';"
        "        font : 20px;"
        "    }"
        ""
        "    QScrollArea{"
        "        border:none;"
        "        background : white;"
        "    }"
        ""
        "    QScrollBar::add-line:vertical {"
        "        border:none;"
        "        background: white;"
        "        height: 0px;"
        "        subcontrol-position: bottom;"
        "        subcontrol-origin: margin;"
        "    }"
        ""
        "    QScrollBar::sub-line:vertical {"
        "        border:none;"
        "        background: white;"
        "        height: 0px;"
        "        subcontrol-position: top;"
        "        subcontrol-origin: margin;"
        "    }"
        ""
        "    QScrollBar::add-page:vertical {"
        "        background: white;"
        "    }"
        ""
        "    QScrollBar::sub-page:vertical {"
        "        background: white;"
        "    }"
        ""
        "    QTableView::item::selected {"
        "        border-top: 2px solid #00acb0;"
        "        border-bottom: 2px solid #00acb0;"
        "        color:black;"
        "    }"
        "    QTextEdit {"
        "        background : white;"
        "    }"
        )
        self.show_buttons() #appel de la focntion show_buttons pour afficher la page de selection de ficheir

    def connexion(self, type_num): #fonction connexion(self, type de rappel effetuer ( permco, rdv, parcours) type_num: int)
        self.clear_widgets() #supprime les widget présent dans le container
        self.set_bg() #ajoute l'image emmaus connect
        self.data = openpyxl.load_workbook("./database/db.xlsx") #ouverture de la base de donnée

        self.ip_input = QLineEdit(self.container) #création du champo de texte recevant l'adresse ip
        self.ip_input.setPlaceholderText("Entrez l'adresse IP du téléphone") #ajout d'un text de fond en cas de champs vide
        self.ip_input.setStyleSheet("background-color: white;") #couleur du champ de text blanc
        self.ip_input.setGeometry(350, 250, 500, 50)  # Position/taille du champs de texte dans le container
        self.ip_input.setFont(
            QFont("Fira Sans", 14)
        )  # Agrandir la police du champ de texte
        self.ip_input.setText(self.data.active.cell(row=4, column=1).value) #définit le texte par la dernière adresse ip saisi et enregistré dans la base de donnée
        self.ip_input.setAlignment(Qt.AlignCenter) #texte centré dans le champs de texte
        self.ip_input.show() #afficher l'élément

        self.send_button = QPushButton("Envoyer SMS", self.container) # création du bouton Envoyer SMS
        self.send_button.setGeometry(500, 350, 200, 50)  # Position/taille du bouton
        self.send_button.clicked.connect(self.connect_and_send_sms) #connexion du bouton à la fonction connect_and_send_sms
        self.send_button.show() #afficher l'élément 

        self.quit_button = QPushButton("Quitter", self.container) #création du bouton quitter
        self.quit_button.setGeometry(950, 25, 200, 50)  # Position/taille du bouton
        self.quit_button.clicked.connect(self.quit) #connexion du bouton à la fonction quit pour fermer la page
        self.quit_button.show() #afficher l'élément

        self.back_button = QPushButton("Retour", self.container) #création du bouton retour
        self.back_button.setGeometry(725, 25, 200, 50)  # Position/taille du bouton
        self.back_button.clicked.connect(partial(self.mess, type_num))#connexion du bouton retour à fonction mess avec comme argument le type de rappel(int)
        self.back_button.show()#afficher l'élément
  
    def quit(self): #fonction quit(self) qui permet de fermet l'application
        QApplication.instance().quit()

    def connect_and_send_sms(self): #fonction connect_and_send_sms(self) qui lance le trhad de connexion airemere en lui passant les arguments ip et type_num
        ip_address = self.ip_input.text() #récupération de la valeur du champ ip ( adress ip pour la connexion)
        sheet = self.selected_file #récupération de la feuille xlsx ouverte précédement
        self.data.active.cell(row=4, column=1).value = ip_address #ajout de l'adresse ip renseigner dans la base de données
        self.data.save("./database/db.xlsx") #sauvegarde des modificationd e la base de données

        if not ip_address: #si aucune adresse ip renseigner
            QMessageBox.critical(self, "Erreur", "Veuillez entrer une adresse IP.") #message d'erreur
            return 

        self.connection_thread = AirmoreConnectionThread(ip_address, sheet) #stock le retour du thread de connexion (succes/error)
        self.connection_thread.connection_success.connect( #connect le signal de réussite à la fonction d'affichage de réussite
            self.show_connection_success_message
        )
        self.connection_thread.connection_error.connect( #connect le signal d'erreur à l'affichage du signal d'erreur
            self.show_connection_error_message
        )
        self.connection_thread.start() #lance le thread de connexion

        self.progress_dialog = QProgressDialog(
            "Accepter la connexion sur le téléphone...", None, 0, 0 #crée une boite de dialogue de chargement pour attendre la validation utilisateur de la connexion
        )
        self.progress_dialog.setWindowModality(Qt.WindowModal) #paramètre de la boite de dialogue
        self.progress_dialog.setWindowTitle("Connexion en cours") #titre de la boite de dialogue
        self.progress_dialog.setCancelButton(None) #supression du bouton annuler
        self.progress_dialog.show() #afficher l'élément

    def show_connection_success_message(self): #fonction d'affichage du message en cas de succès de la connexion
        self.progress_dialog.close() #fermeture de la boite de dialogue de chargement
        self.clear_widgets() #su^pression des élément du container

        msg = QMessageBox()#création d'un message de réussite de connexion
        msg.setText("La connexion avec le téléphone a été établie avec succès. Tous les messages seront envoyés") #texte du message
        msg.setWindowTitle("Connexion réussie") #titre
        msg.setIcon(QMessageBox.Information) #type de box message
        msg.setStyleSheet("background-color: rgb(255, 255, 255)") #définit la couleur de fond
        msg.exec_() #affciher le message

        self.show_buttons() #appeler la fonction show_bouton pour retourner au choix de type de rappel

    def show_connection_error_message(self, error_message): #fonction d'affiche du message en cas d'erreur
        self.progress_dialog.close() #fermer la boite de dialogue de chargement

        msg = QMessageBox() #création du message
        msg.critical(
            self.container,
            "Erreur",
            f"Impossible de se connecter avec l'adresse IP {self.ip_input.text()}: {error_message}", #texte du message avec ajout du retour de l'erreur
        )
        msg.setStyleSheet("QWidget { background-color: white; }") #couleur de fond du message

    def clear_widgets(self): #fonction clear_widgets(self) qui permet de supprimer tout les widget présent dans le container afin d'en afficher de nouveau
        for sub_widget in self.container.findChildren(QWidget): #pour chaque enfant du container
            sub_widget.deleteLater() #le supprimer

    def browse_file(self): #fonction de parcours des fichier .xlsx

        file_path, _ = QFileDialog.getOpenFileName(
            self, "Sélectionner un fichier", "", "Excel (*.xlsx)"  # Ouvrir une boîte de dialogue pour sélectionner un fichier
        )

        if file_path: #si fichier sélectionné
            self.selected_file = file_path

            msg = QMessageBox()                                 #        ""
            msg.setText(f"Le fichier sélectionné est : {file_path}")
            msg.setWindowTitle("Fichier sélectionné")       #        message de succès de sélection d'un fichier
            msg.setIcon(QMessageBox.Information)
            msg.setStyleSheet("background-color: rgb(255, 255, 255)")
            msg.exec_() #                                                ""

            self.file_input.setText(file_path) # remplis le champs de text avec le chemin du fichier
            self.file_input.setReadOnly(False) #rend le chemin editable dans le champs de texte

    def show_buttons(self): #fonction principal de sélection de fichier et choix du type de rappel
        self.clear_widgets() #supprimer les élément du container
        self.set_bg() #ajout de l'image emmaus connect

        self.permco_button = QPushButton("PermCo", self.container) #création du bouton "permco"
        self.permco_button.setGeometry(500, 350, 200, 50) #place/taille du bouton
        self.permco_button.clicked.connect(partial(self.handle, 1))#connexion du bouton à la fonction self.handle en passant l'argument "1"
        self.permco_button.show()#afficher l'élément

        self.file_input = QLineEdit(self.container) #création du champ de text du chemin du fichier xlsx
        self.file_input.setPlaceholderText("Chemin du rapport excel") #texte de fond lorsque champ vide
        self.file_input.setStyleSheet("background-color: white;") #couleur du champ en blanc
        self.file_input.setGeometry(350, 250, 500, 50)  #place/taille du champs 
        self.file_input.setFont(QFont("Fira Sans", 14)) #police d'écriture du champ
        self.file_input.setReadOnly(True) #bloque l'édition du texte tant que non remplis
        self.file_input.show() #afficher l'élément

        try: #charger le fichier si selectionner précédement ( back bouton) et reset z colonnes
            self.file_input.setText(self.selected_file)
            self.file_input.setReadOnly(False)
            for row_index in range(
                    1, self.file_data.active.max_row + 1
                ):
                self.file_data.active.cell(row=row_index, column=26).value = (
                            0  # remise a zero de la cellule z
                        )
            self.file_data.save(self.selected_file) #enregistrement des modifs
        except Exception as e:
            print("pas de fichier charger", e)

        self.parcours_button = QPushButton("Parcours", self.container)#création du bouton parcours
        self.parcours_button.setGeometry(500, 425, 200, 50) #place/taille du bouton
        self.parcours_button.clicked.connect(partial(self.handle, 2)) #connexion du bouton à la fonction handle avec l'agument "2"
        self.parcours_button.show() #afficher l'élément

        self.rendezvous_button = QPushButton("Rendez-vous", self.container) #création du bouton rendez-vous
        self.rendezvous_button.setGeometry(500, 500, 200, 50)#place /taille du bouton
        self.rendezvous_button.clicked.connect(partial(self.handle, 3))#connexion du bouton à la fonction handle avec l'argument "3"
        self.rendezvous_button.show()#afficher l'élément

        self.browse_button = QPushButton("Choisir un Fichier", self.container)#création du bouton parcourir pour le choix de fichier
        self.browse_button.setGeometry(350, 150, 500, 50)  # Position/taille du bouton
        self.browse_button.clicked.connect(self.browse_file) #connexion du bouton avec la fonction browse_file
        self.browse_button.show()#afficher l'élément

        self.quit_button = QPushButton("Quitter", self.container)#création du bouton quitter
        self.quit_button.setGeometry(950, 25, 200, 50)  # Position/taille du bouton
        self.quit_button.clicked.connect(self.quit) #connexion du bouton avec la fonction quit()
        self.quit_button.show()#afficher l'élément

    def is_file_valid(self, file_path):#fonction qui permet d'ouvrir le fichier et gérer les erreurs

        try:
            wb = openpyxl.load_workbook(file_path)# Tentative d'ouverture du fichier avec openpyxl
            self.file_data = wb # Stocker les données du fichier 
            return True
        
        except Exception as e:
            self.file_data = None# Réinitialiser l'attribut de données du fichier
            return e #retourne l'erreur

    def handle(self, type_num): #fonction d'affichage des données du fichier xlsx sous forme de tableau et selection des contacts pour envoi du sms

        if not hasattr(self, "selected_file") or not self.selected_file: #vérifie qu'un fichier à bien été sélectionné
            QMessageBox.warning(
                self, "Avertissement", "Veuillez d'abord sélectionner un fichier." #message si non
            )
            return

        if self.is_file_valid(self.selected_file) != True: #vérifie si le fichier se charge correctement
            QMessageBox.warning(
                self,
                "Avertissement",
                f"Impossible de charger le fichier : {self.is_file_valid(self.selected_file)}", #message en cas d'erreur
            )
            return
        
        sheet = self.file_data.active #récupération et stockage de la feuille active 
        num_rows = sheet.max_row #nombre de ligne de données dans la feuille
        displayed_row_index = 0 #initialise l'index xlsx
        num_displayed_rows = 0 #initialie l'index du tableau de l'app

        type_text = "" #init var type_text
        if type_num == 1: #si argument "1" passer
            type_text = "Permanence Connectée" # modif de "type_text"
        elif type_num == 2: #si argument "2" passer
            type_text = "Parcours d'initiation"  
        elif type_num == 3:#si argument "3" passer
            type_text = "RDV bénéficiaire"

        
        try:
            for row_index in range(1, num_rows + 1): #pour chaque lignes dans le fichier excel
                type_rdv = sheet.cell(
                    row=row_index, column=10 #recupere le type de rdv de la ligne
                ).value  # Colonne J
                #sheet.cell(row=row_index, column=26).value = 0 #inittialise la colonne z à 0
                if type_text in type_rdv: #si type de rdv == le type choisi
                    num_displayed_rows += 1 #compter +1
        except:
            QMessageBox.warning(
                self,
                "Avertissement",
                f"Impossible de charger le fichier : {self.selected_file}", #si problème(s) dans la lecture du fichier message d'erreur et retour
            )
            return
        
        self.clear_widgets() #effacer les widgets présent dans le container
        self.set_bg()#ajouter la photo emmaus connect

        self.table_widget = QTableWidget(self.container) #création du tableau de valeur
        self.table_widget.setGeometry(40, 100, 1120, 650) #taile/place du tableau dans le conatiner
        
        headers = [ # Définir les entêtes de colonnes
            "Nom complet",
            "Date",
            "Heure",
            "Numéro Tel",
            "Observations",
            "Statut",
            "Rappel",
        ]

        self.table_widget.setColumnCount(len(headers)) #définit le nombre de colonnes
        self.table_widget.setHorizontalHeaderLabels(headers) #ajoute les entêtes
        self.table_widget.setColumnWidth(0, 200) #                      ""
        self.table_widget.setColumnWidth(1, 150)
        self.table_widget.setColumnWidth(2, 100)
        self.table_widget.setColumnWidth(3, 200)#                    définit la largeur de chaques colonnes
        self.table_widget.setColumnWidth(4, 200)
        self.table_widget.setColumnWidth(5, 120)
        self.table_widget.setColumnWidth(6, 90)  #                      ""

        self.table_widget.setRowCount(num_displayed_rows) # définit le nombre de lignes dans le tableau (en fonction du nombre de contact)

        for row_index in range(1, num_rows + 1): # pour chaque ligne du fichier xlsx
            type_rdv = sheet.cell(row=row_index, column=10).value  #recupere le type de rdv
            if type_text in type_rdv: # si type de rdv == type de rdv choisi
                
                nom_complet = QTableWidgetItem(#                                                 ""
                    str(sheet.cell(row=row_index, column=4).value)
                ) 
                date = str(sheet.cell(row=row_index, column=2).value)[:10]
                heure = str(sheet.cell(row=row_index, column=2).value)[11:]  
                numero_tel = sheet.cell(row=row_index, column=5).value              # Récupérer les valeurs des colonnes spécifiées
                statut_z = sheet.cell(row=row_index, column=26).value
                statut_participant = QTableWidgetItem(
                    str(sheet.cell(row=row_index, column=7).value)
                )
                if type_num == 3: #si type choisi == rendez-vous bénéficiaire (inscription/vente)
                    if str(sheet.cell(row=row_index, column=8).value) == "None": #si observation dans la ligne == None
                        observation = QTableWidgetItem(
                            str(sheet.cell(row=row_index, column=1).value) #ecrire le type de rdv benef dans observation
                        )
                    else:
                        observation = QTableWidgetItem(
                            str(sheet.cell(row=row_index, column=1).value)
                            + ": "                                          #sinon ecrire le type de rdv benef + observation indiquée
                            + str(sheet.cell(row=row_index, column=8).value)
                        )
                else: #sinon
                    if str(sheet.cell(row=row_index, column=8).value) == "None": #si observation = None laissé la case observation vide
                        observation = QTableWidgetItem("")
                    else:
                        observation = QTableWidgetItem(
                            str(sheet.cell(row=row_index, column=8).value) #sinon remplir avec l'observation indiquée
                        ) 

                                                                 #                         ""

                self.table_widget.setItem(displayed_row_index, 0, nom_complet) #remplis la case nom_complet avec la valeur du fichier
                nom_complet.setTextAlignment(Qt.AlignCenter) #alignement du texte au centre
                nom_complet.setFont(QFont("Fira Sans", 12)) #police du texte

                self.table_widget.setItem(displayed_row_index, 4, observation) #remplis la case observation avec la valeur du fichier
                observation.setFont(QFont("Fira Sans", 12)) #police du texte

                self.table_widget.setItem(displayed_row_index, 5, statut_participant) #remplis la case statut_participant avec la valeur du fichier
                statut_participant.setTextAlignment(Qt.AlignCenter) #alignement du texte au centre
                statut_participant.setFont(QFont("Fira Sans", 12)) #police du texte

                checkbox_widget = QWidget() #création d'un widget container pour la case à cocher
                checkbox_layout = QHBoxLayout() #création d'une grille de placement pour la case à cocher
                checkbox = QCheckBox() #création de la case à cocher
                checkbox.setChecked(bool(statut_z)) #bascule l'état de la case a cocher en focntion de la valeur de la colonne z du fichier
                checkbox.setProperty("id", str(row_index)) #ajout de l'index de la ligne dans le fichier comme identifiant de la case à cocher
                checkbox_layout.addWidget(checkbox) #ajout de la case à cocher à la grille de placement
                checkbox_layout.setAlignment(Qt.AlignCenter) #place au centre la case à cocher
                checkbox_widget.setLayout(checkbox_layout) #ajout de la grille de placement au widget
                self.table_widget.setCellWidget(displayed_row_index, 6, checkbox_widget) #ajout du widget de la case à cocher au tableau

                item = QTableWidgetItem(str(date)) #définit le widget de la date
                if not self.is_valid_date(date): #si date non valide
                    item.setForeground(QColor(Qt.red)) #affichage en rouge
                    checkbox_widget.setEnabled(False) #désactive la case à cocher
                item.setTextAlignment(Qt.AlignCenter) #aligne le texte au centre de la cellule
                item.setFont(QFont("Fira Sans", 12)) #police du texte
                self.table_widget.setItem(displayed_row_index, 1, item) #ajoute la valeur au tableau

                item = QTableWidgetItem(str(heure)) #définit le widget de l'heure
                if not self.is_valid_hour(heure):#si heure non valide
                    item.setForeground(QColor(Qt.red))#affichage en rouge
                    checkbox_widget.setEnabled(False) #désactive la case à cocher
                item.setTextAlignment(Qt.AlignCenter)#aligne le texte au centre de la cellule
                item.setFont(QFont("Fira Sans", 12))#police du texte
                self.table_widget.setItem(displayed_row_index, 2, item)#ajoute la valeur au tableau

                item = QTableWidgetItem(str(numero_tel))#définit le widget du numéro de tel
                if not self.is_valid_phone_number(numero_tel):#si numéro de tel non valide
                    item.setForeground(QColor(Qt.red))#affichage en rouge
                    checkbox_widget.setEnabled(False)#désactive la case à cocher
                item.setTextAlignment(Qt.AlignCenter)#aligne le texte au centre de la cellule
                item.setFont(QFont("Fira Sans", 12))#police du texte
                self.table_widget.setItem(displayed_row_index, 3, item)#ajoute la valeur au tableau

                self.table_widget.setItemDelegate(FontDelegate()) #définit la police lorsque le texte est en cours de modification
                displayed_row_index += 1 #ajoute 1 à l'index pour passer à la ligne suivante

        self.table_widget.cellChanged.connect(self.set_enabled) #connect le changement d'une cellule à la fonction set_enabled qui verifie si les informations sont valides

        self.global_checkbox = QCheckBox(
            "Sélectionner/Désélectionner Tout", self.container #création de la case à cocher Sélectionner/Désélectionner Tout
        )
        self.global_checkbox.setGeometry(50, 750, 280, 50)  # Position/place de la case
        self.global_checkbox.setStyleSheet("color: white;") #couleur du texte
        self.global_checkbox.setChecked(False) #initialise l'état de la case à decocher
        self.global_checkbox.stateChanged.connect(self.check_uncheck_all) #connect la case à cocher avec la fonction check_uncheck_all
        self.global_checkbox.show() #afficher l'élement

        self.table_widget.resizeRowsToContents() #redéfinit la hauteur des lignes à leur contenus

        self.validate_button = QPushButton("Valider", self.container) #création du bouton valider
        self.validate_button.setGeometry(500, 760, 200, 50) #position/taille du bouton

        self.validate_button.clicked.connect(
            partial(self.save_checked_values, type_num) #connect le bouton avec la fonction save_checked_value en passant pour argument le type de rdv choisi
        )
        self.validate_button.show() #afficher l'élement

        self.quit_button = QPushButton("Quitter", self.container) #création du bouton quitter
        self.quit_button.setGeometry(950, 25, 200, 50)  # Position/taille du bouton
        self.quit_button.clicked.connect(self.quit) #connexion du bouton avec la fonction quit()
        self.quit_button.show() #afficher l'element

        self.back_button = QPushButton("Retour", self.container) #création du bouton retour
        self.back_button.setGeometry(725, 25, 200, 50)  # Position/taille du bouton
        self.back_button.clicked.connect(self.show_buttons) #connexion du bouton à la fonction précédente show_boutons
        self.back_button.show()#afficher l'élément

        self.table_widget.show() #afficher le tableau

    def check_uncheck_all(self, state): # Cocher ou décocher toutes les cases à cocher de la colonne en fonction de l'état de la case à cocher de l'en-tête

        for row_index in range(self.table_widget.rowCount()): #pour chaque ligne du tableau
            checkbox_widget = self.table_widget.cellWidget(row_index, 6) #récupere le widget de la case à cocher
            if (
                self.check_phone_number(row_index, 3)
                and self.check_date(row_index, 1)    #si toutes les valeurs sont valides
                and self.check_hour(row_index, 2)
            ):
                if checkbox_widget is not None:    #si le widget de la case existe
                    checkbox = checkbox_widget.findChild(QCheckBox)  #récupere la case à cocher
                    if checkbox is not None: #si case à cocher existe
                        checkbox.setChecked(state == Qt.Checked) # déifnit la case comme cocher/deccocher

    def update_status_z(self, state, row, row_xls, type_num):# Mettre à jour l'état de la colonne Z dans le fichier xlsx en fonction de l'état de la case à cocher
        #update status_z(état de la case, index de la ligne du tableu, index de la ligne fichier, type de rdv)

        sheet = self.file_data.active  # Récupérer la feuille active

        if int(state) == 1: #si case cocher
            sheet.cell(row=row_xls, column=26).value = int(type_num) #remplit la colonne z par le type de rdv (format num)
        else:
            sheet.cell(row=row_xls, column=26).value = int(state) #sinon met la colonne z à 0

                                                                            #remplit le fichier par les valeurs modifiées du tableau
        sheet.cell(row=row_xls, column=4).value = self.table_widget.item( #                       ""
            row, 0
        ).text()
        sheet.cell(row=row_xls, column=2).value = (
            self.table_widget.item(row, 1).text()
            + " "
            + self.table_widget.item(row, 2).text()
        )
        sheet.cell(row=row_xls, column=5).value = self.table_widget.item(row, 3).text()
        sheet.cell(row=row_xls, column=8).value = self.table_widget.item(
            row, 4
        ).text()  # Colonne E
        sheet.cell(row=row_xls, column=7).value = self.table_widget.item(row, 5).text()
                                                                        #                         ""

        self.file_data.save(self.selected_file) #sauvegarde les modifications dans le fichier

    def save_checked_values(self, type_num): #fonction permettant l'enregistrement des modifications et le compte du nombre de case coché

        if self.file_data is None:
            QMessageBox.warning(
                self, "Avertissement", "Veuillez d'abord sélectionner un fichier." #vérifie qu'un fichier est bien sélectionné pour eviter les erreurs
            )
            return
        
        count = 0 #initialise le compte de case cocher à 0

        for row_index in range(self.table_widget.rowCount()): # Parcourir toutes les lignes du tableau
            checkbox_widget = self.table_widget.cellWidget(row_index, 6) # reupere le widget de case à cocher
            if checkbox_widget is not None: #si widget existe
                checkbox = checkbox_widget.findChild(QCheckBox) #récupere la case à cocher
                if checkbox is not None: #si case à cocher existe
                    checked = checkbox.isChecked()# Récupérer l'état de la case à cocher
                    if checked: #si case cocher
                        count += 1 #compter +1
                    

                    self.update_status_z(
                        checked, row_index, int(checkbox.property("id")), type_num # Mettre à jour la valeur dans la colonne Z du fichier xlsx (update status_z(état de la case, index de la ligne du tableu, index de la ligne fichier, type de rdv))
                    ) 

        if count > 0: # Si nombre de case cocher > 0
            QMessageBox.information(
                self,
                "Succès",
                "Les valeurs ont bien été enregistrés, " #afficher message de succès et nombre de case cocher
                + str(count)
                + " message-s ser-a-ont envoyés",
            )
            self.mess(type_num)
        else:
            QMessageBox.warning(
                self, "Avertissement", "Veuillez sélectionner des contacts" #sinon afficher message d'erreur pour demander de sélectionné des contacts
            )

    def check_phone_number(self, row, column): #verifie que le numero de téléphone est dans un format valide

        item = self.table_widget.item(row, 3) #recupere le widget du numéro de téléphone du tableau

        checkbox_widget = self.table_widget.cellWidget(item.row(), 6) #récupere le widget de la case à cocher
        checkbox = checkbox_widget.findChild(
            QCheckBox #récupere la case à cocher de la ligne
        ) 

        phone_number = item.text() #récupere le numéro de tléphone dans le widget

        if not self.is_valid_phone_number(phone_number): #si format invalide ( is_valid_phone_number return False)
            item.setForeground(QColor(Qt.red)) #texte en rouge
            checkbox.setChecked(False) #décoche la case à cocher
            checkbox_widget.setEnabled(False) #désactive la case à cocher
            return False #retour Faux
        else:
            item.setForeground(QColor(Qt.black)) #sinon met le texte en noir
            return True #retour Vrai

    def check_date(self, row, column): #verifie que la date est valide 

        item = self.table_widget.item(row, 1) #recupere le widget de la date dans le tableau

        checkbox_widget = self.table_widget.cellWidget(item.row(), 6)#récupere le widget de la case à cocher
        checkbox = checkbox_widget.findChild(
            QCheckBox   #récupere la case à cocher de la ligne
        )

        date = item.text() #récupere la date dans le widget 

        if not self.is_valid_date(date): #si date non valide 
            item.setForeground(QColor(Qt.red)) #couleur du texte en rouge
            checkbox.setChecked(False) #décoche la case à cocher de la ligne
            checkbox_widget.setEnabled(False) #désactive la case à cocher
            return False #retourne faux
        else:
            item.setForeground(QColor(Qt.black)) #sinon met le texte en noir
            return True #retourne Vraie

    def check_hour(self, row, column): #verifie si le format de l'heure est valide

        item = self.table_widget.item(row, 2) #recupere le widget de l'heure du tableau

        checkbox_widget = self.table_widget.cellWidget(item.row(), 6)#récupere le widget de la case à cocher
        checkbox = checkbox_widget.findChild(
            QCheckBox    #récupere la case à cocher de la ligne
        )

        hour = item.text() #récupere l'heure dans le widget

        if not self.is_valid_hour(hour): #si heure non valide 
            item.setForeground(QColor(Qt.red))#couleur du texte en rouge
            checkbox.setChecked(False)#décoche la case à cocher de la ligne
            checkbox_widget.setEnabled(False)#désactive la case à cocher
            return False#retourne faux
        else:
            item.setForeground(QColor(Qt.black))#sinon met le texte en noir
            return True#retourne Vraie

    def set_enabled(self, row, column): #lance les fonctions qui teste si les formats sont correctes pour activer la case à cocher

        checkbox_widget = self.table_widget.cellWidget(row, 6) #récupere le widget de la case à cocher

        if (
            self.check_phone_number(row, column) #si toute les valeurs sont valides
            and self.check_date(row, column)
            and self.check_hour(row, column)
        ):
            checkbox_widget.setEnabled(True) #active la case à cocher
        else:#sinon 
            self.check_date(row, column) #verifie tout de même la date et l'heure
            self.check_hour(row, column)

    def is_valid_phone_number(self, phone_number):
        # Vérifier si le numéro de téléphone est valide selon l'expression régulière

        return bool(re.match(r"^(?:\D*\d){0}\D*0(?:\D*\d){9}\D*$", str(phone_number)))

    def is_valid_date(self, date):
        # Vérifier si la date est valide selon l'expression régulière

        return bool(
            re.match(r"^(0[1-9]|[12][0-9]|3[01])/(0[1-9]|1[0-2])/\d{4}$", str(date))
        )

    def is_valid_hour(self, hour):
        # Vérifier si l'heure est valide selon l'expression régulière

        return bool(re.match(r"^([01][0-9]|2[0-3]):[0-5][0-9]$", str(hour)))

    def mess(self, type_num): #permet l'edition du message avant envoi

        self.clear_widgets() #supprime les widgets présent dans le container
        self.set_bg() #ajoute l'image emmaus connect

        self.quit_button = QPushButton("Quitter", self.container) #création du bouton quitter 
        self.quit_button.setGeometry(950, 25, 200, 50)  # Position/taille du bouton
        self.quit_button.clicked.connect(self.quit) #connexion du bouton avec la fonction quit()
        self.quit_button.show() #afficher l'élément

        self.back_button = QPushButton("Retour", self.container) #création du bouton retour
        self.back_button.setGeometry(725, 25, 200, 50)  # Position / taille du bouton
        self.back_button.clicked.connect(partial(self.handle, type_num)) #connexion du bouton retour avec la focntion handle prenant en parametre le type de rdv format num
        self.back_button.show() #afficher l'élélement

        self.send_button = QPushButton("Envoyer", self.container) #création du bouton envoyer
        self.send_button.setGeometry(500, 700, 200, 50)  # Position/taille du bouton
        self.send_button.clicked.connect(partial(self.save_db_changed, type_num)) #connexion du bouton envoyer avec la fonction save_db_changed avec pour argument le type de rdv choisi
        self.send_button.show() #afficher l'élément

        self.text_edit = QTextEdit(self.container) #création de la zone d'édition de texte 
        self.text_edit.setPlaceholderText("Entrez votre message ici...") #définit le text si zone vide
        self.text_edit.setText(str(self.db.active.cell(type_num, 1).value)) #remplis la zone de texte par le message stocké dans la base de donnée
        #self.onTextChanged #appel la fonction de verification de keyword une première fois
        self.text_edit.setGeometry(300, 150, 600, 450) #position/taille de la zone de texte
        self.text_edit.setFont(QFont("Fira Sans", 20)) #police du texte
        self.text_edit.textChanged.connect(self.onTextChanged) #connect les changement du texte avec la fonction onTextChanged
        self.highlightKeywords() #met en srubrillance les keywords
        self.text_edit.show() #affiche l'élément

    def onTextChanged(self): #fonction de temporisation pour ne pas surcharger d'exécution la fonction de surbrillance 
        if not self.highlighting: # si self.highlighting n'est pas en cours 
            self.highlighting = True
            self.highlightKeywords() #l'exécuter
            self.highlighting = False

    def save_db_changed(self, type_num): #fonction de sauvegardes des changements des messages de la db
        self.db.active.cell(1, 1).value = self.text_edit.toPlainText() #récupere le message dans la zone d'édition de texte et remplace l'encienne version dans la DB
        self.db.save("./database/db.xlsx") #sauvegarde les modifications dans la base de données
        self.connexion(type_num) #appel la fonction connexion pour continuer

    def highlightKeywords(self):

        text = self.text_edit.toPlainText()# Récupérer le texte actuel du QTextEdit
        keywords = ["[heure]", "[date]", "[nom]", "[prenom]", "[tel]", "[jour_semaine]", "[mois]", "[annee]", "[jour_num]"]# Liste de mots-clés à rechercher
        
        red_format = QTextCharFormat()# Créer un objet QTextCharFormat pour le format de mise en évidence (cyan)
        red_format.setForeground(QColor("#00acb0")) 

        cursor = self.text_edit.textCursor()# Créer un objet QTextCursor

        cursor.select(QTextCursor.Document) # Supprimer tout formatage existant
        cursor.setCharFormat(QTextCharFormat())

        # Parcourir la liste de mots-clés et mettre en surbrillance chaque occurrences dans le texte

        for keyword in keywords: 
            index = text.find(keyword) 
            while index != -1: 
                cursor.setPosition(index) 
                cursor.movePosition(
                    QTextCursor.Right, QTextCursor.KeepAnchor, len(keyword)
                )
                cursor.setCharFormat(red_format)
                index = text.find(keyword, index + len(keyword))


if __name__ == "__main__": #fonction principal :main
    app = QApplication(sys.argv) #creation de l'application pyqt5
    dir_ = QDir("Fira Sans") #creation d'une police
    _id = QFontDatabase.addApplicationFont("Fira Sans/FiraSans-Regular.ttf") #chemin de la police
    window = MainWindow() #creation de la fenetre principal
    window.show() #afficher l'élément
    sys.exit(app.exec_()) #fermeture si clique sur la croix
  